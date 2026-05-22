import sqlite3
import shutil
import os
import zipfile
import re
from datetime import datetime, timedelta
import openpyxl
from db_manager import DB_NAME, DATA_LABELS

def clean_external_links_physically(file_path):
    """
    오류 팝업창의 근본 원인이 되는 엑셀 내부의 외부 링크 파편(externalLinks)들을
    Zipfile 레이어에서 완벽하게 박멸하는 물리 세척기 함수입니다.
    """
    temp_file_path = file_path + ".tmp"
    try:
        with zipfile.ZipFile(file_path, 'r') as zin:
            with zipfile.ZipFile(temp_file_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    filename = item.filename
                    if "externalLinks/" in filename:
                        continue
                    if filename == "xl/_rels/workbook.xml.rels":
                        data = zin.read(filename).decode('utf-8')
                        data_clean = re.sub(r'<Relationship[^>]*Type="[^"]*externalLink"[^>]*/>', '', data)
                        zout.writestr(filename, data_clean.encode('utf-8'))
                        continue
                    if filename == "xl/workbook.xml":
                        data = zin.read(filename).decode('utf-8')
                        data_clean = re.sub(r'<externalReferences>.*?</externalReferences>', '', data, flags=re.DOTALL)
                        zout.writestr(filename, data_clean.encode('utf-8'))
                        continue
                    zout.writestr(item, zin.read(filename))
        os.remove(file_path)
        os.rename(temp_file_path, file_path)
    except Exception as e:
        if os.path.exists(temp_file_path): os.remove(temp_file_path)
        print(f"[경고] 외부 링크 물리 청소 중 예외 발생: {e}")

def generate_excel_report(selected_date):
    """
    22행의 start("KEP_P_kWh"), end("KEP_P_kWh") 서식을 정확히 매칭하여 
    실측 데이터를 주입하고 테이블 헤더(24, 25행)를 원본 그대로 보존하는 핵심 엔진입니다.
    """
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    template_file = os.path.join(BASE_DIR, "template_전기실_운영일지.xlsx")
    output_file = os.path.join(BASE_DIR, f"{selected_date}_전기실_운영일지.xlsx")
    
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"템플릿 파일 [{template_file}]이 경로에 존재하지 않습니다.")
    
    # 1. 템플릿 안전 복사 및 워크북 로드
    shutil.copy(template_file, output_file)
    wb = openpyxl.load_workbook(output_file, data_only=False)
    
    # 오픈픽셀 자체 링크 버그 예방책 마련
    if hasattr(wb, 'external_link_refs'): wb.external_link_refs = []
    if hasattr(wb, '_external_links'): wb._external_links = []

    dt = datetime.strptime(selected_date, "%Y-%m-%d")
    date_str_formatted = f"{dt.year}년 {dt.month:02d}월 {dt.day:02d}일"
    
    # 시트 명칭 동적 보정 및 할당
    ws_summary = None
    ws_detail = None
    for sheet in wb.worksheets:
        if "운영현황" in sheet.title or "운전현황" in sheet.title:
            ws_summary = sheet
            sheet.title = f"{selected_date}_전력설비_운전현황"
        elif "상세내역" in sheet.title:
            ws_detail = sheet
            sheet.title = f"{selected_date}_상세내역"
            
    if not ws_summary: ws_summary = wb.worksheets[0]
    if not ws_detail: ws_detail = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # =========================================================================
    # [파트 1] 전력설비 운영 현황 분석 자료 - 최고/최저값 덮어쓰기 (상단 구역)
    # =========================================================================
    c.execute('SELECT * FROM daily_extremes WHERE log_date = ?', (selected_date,))
    rows_extreme = c.fetchall()
    c.execute('PRAGMA table_info(daily_extremes)')
    columns_extreme = [col[1] for col in c.fetchall()]
    
    extremes_dict = {}
    for r in rows_extreme:
        ext_type = r[columns_extreme.index('extreme_type')]
        for col_name in DATA_LABELS:
            if col_name in columns_extreme:
                val = r[columns_extreme.index(col_name)]
                extremes_dict[(ext_type, col_name)] = round(val, 1) if val is not None else "-"

    for row in ws_summary.iter_rows(max_row=20):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_strip = cell.value.strip().replace(" ", "")
                if "일시:" in val_strip or ("일시" in val_strip and ":" in val_strip):
                    cell.value = f"일시 : {date_str_formatted}"
                elif "max(" in val_strip:
                    label = cell.value.split('"')[1].strip() if '"' in cell.value else cell.value.replace("max(", "").replace(")", "").strip()
                    cell.value = extremes_dict.get(("MAX", label), "-")
                elif "min(" in val_strip:
                    label = cell.value.split('"')[1].strip() if '"' in cell.value else cell.value.replace("min(", "").replace(")", "").strip()
                    cell.value = extremes_dict.get(("MIN", label), "-")

        
    # =========================================================================
    # 💡 [파트 2 수정 완료] 22행 고정 정밀 타격 및 영문 식별자 매핑 처리
    # =========================================================================
    # 1) 월시작 누적 전력량 추적 (당월 1일부터 선택일까지 데이터 중 최초 누적값)
    start_of_month = f"{dt.year}-{dt.month:02d}-01"
    c.execute('''
        SELECT KEP_P_kWh FROM raw_data 
        WHERE log_date >= ? AND log_date <= ? AND KEP_P_kWh IS NOT NULL 
        ORDER BY log_date ASC, log_time ASC LIMIT 1
    ''', (start_of_month, selected_date))
    month_start_res = c.fetchone()
    month_start_val = round(month_start_res[0], 1) if month_start_res and month_start_res[0] is not None else "-"

    # 2) 당일 최종 전력량 추적 (보고서 지정 날짜의 최고 누적 전력량 값)
    c.execute('SELECT MAX(KEP_P_kWh) FROM raw_data WHERE log_date = ?', (selected_date,))
    row_end_res = c.fetchone()
    day_end_val = round(row_end_res[0], 1) if row_end_res and row_end_res[0] is not None else "-"

    # 3) 22행 전체 셀을 수평으로 정밀 스캔하여 데이터 교체 및 수식 주입
    for cell in ws_summary[22]:
        if cell.value and isinstance(cell.value, str):
            text_clean = cell.value.strip().replace(" ", "")
            
            # C22 및 E22 자리 정밀 탐색 치환
            if 'start("KEP_P_kWh")' in text_clean:
                cell.value = month_start_val
            elif 'end("KEP_P_kWh")' in text_clean:
                cell.value = day_end_val
            # G22 자리에 수식이 깨지지 않도록 수식 공식 강제 주입
            elif 'E22-C22' in text_clean:
                cell.value = "=E22-C22"

    # =========================================================================
    # [파트 3] 한전 메인 적산 사용량 하단 요약 테이블 연동 구역 (26행 이하)
    # =========================================================================
    # 전일 종료값(=금일 시작값) 연산용 데이터 조회
    prev_date = (dt - timedelta(days=1)).strftime("%Y-%m-%d")
    c.execute('SELECT MAX(KEP_P_kWh) FROM raw_data WHERE log_date = ?', (prev_date,))
    row_start_res = c.fetchone()
    day_start_val = round(row_start_res[0], 1) if row_start_res and row_start_res[0] is not None else "-"

    today_mwh = day_end_val
    prev_mwh = day_start_val
    diff_mwh = round(today_mwh - prev_mwh, 1) if today_mwh != "-" and prev_mwh != "-" else "-"

    for row in ws_summary.iter_rows(min_row=26):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_strip = cell.value.strip()
                if val_strip == "1469.79": cell.value = today_mwh
                elif val_strip == "1464.06": cell.value = prev_mwh
                elif val_strip == "5.7300000000000182": cell.value = diff_mwh

    # =========================================================================
    # [파트 4] 상세내역 시트 - 시간별 데이터 완벽 주입 (상/하단 24시간 입력 제어)
    # =========================================================================
    for row in ws_detail.iter_rows(max_row=3):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "일자" in cell.value:
                cell.value = f"일자 : {date_str_formatted}"

    avg_selects = ", ".join([f'AVG("{name}")' for name in DATA_LABELS])
    c.execute(f'SELECT CAST(strftime("%H", log_time) AS INTEGER) as hour, {avg_selects} FROM raw_data WHERE log_date = ? GROUP BY hour', (selected_date,))
    hourly_rows = c.fetchall()
    
    hourly_data = {}
    for r in hourly_rows:
        hour = r[0]
        for idx, label in enumerate(DATA_LABELS):
            if r[idx + 1] is not None:
                hourly_data[(hour, label)] = round(r[idx + 1], 1)

    for row_idx, row in enumerate(ws_detail.iter_rows(min_row=4, max_row=60), start=4):
        first_cell_val = str(row[0].value).strip() if row[0].value is not None else ""
        if first_cell_val == "0":
            col_mapping = {}
            for col_idx, cell in enumerate(row):
                if cell.value and isinstance(cell.value, str):
                    clean_label = cell.value.strip().replace('"', '')
                    if clean_label in DATA_LABELS:
                        col_mapping[col_idx] = clean_label
            
            if col_mapping:
                for h in range(24):
                    target_row = row_idx + h
                    ws_detail.cell(row=target_row, column=1).value = h
                    for col_idx, label in col_mapping.items():
                        val = hourly_data.get((h, label), "-")
                        ws_detail.cell(row=target_row, column=col_idx + 1).value = val

    conn.close()
    wb.save(output_file)
    
    # 엑셀 파일 뼈대 속 외부 수식 파편 물리 청소 강제 작동 (오류 경고 완전 제거)
    clean_external_links_physically(output_file)
    print(f"[최종 보완 완료] 정밀 연동 보고서 생성 성공: {output_file}")


def update_power_consumption_section(ws, selected_date):
    """
    운영현황 시트의 21행(일간 전력량) 및 22행(월간 전력량)의 
    텍스트 표식을 실제 DB 데이터로 치환하는 함수
    """
    conn = sqlite3.connect("plc_logging_real.db")
    c = conn.cursor()
    
    # 1. DB에서 금일(selected_date)의 KEP_P_kWh 최고값(마지막) 및 최저값(시작) 조회
    c.execute('''
        SELECT MIN(KEP_P_kWh), MAX(KEP_P_kWh) 
        FROM raw_data 
        WHERE log_date = ?
    ''', (selected_date,))
    res = c.fetchone()
    
    # 데이터가 없을 경우 예외 처리
    today_min_kwh = round(res[0], 1) if res and res[0] is not None else "-"
    today_max_kwh = round(res[1], 1) if res and res[1] is not None else "-"
    
    # 2. 엑셀 시트 전체(혹은 전력량 영역인 20~25행 사이)를 순회하며 텍스트 치환
    for row in ws.iter_rows(min_row=20, max_row=25, min_col=1, max_col=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_text = cell.value.strip()
                
                # [버그 해결] C21 셀의 min("KEP_P_kWh") 표식을 실제 시작 전력량 값으로 치환
                if 'min("KEP_P_kWh")' in cell_text or 'min( "KEP_P_kWh")' in cell_text:
                    cell.value = today_min_kwh
                    
                # [버그 해결] E21 셀의 max("KEP_P_kWh") 표식을 실제 마지막 전력량 값으로 치환
                elif 'max("KEP_P_kWh")' in cell_text or 'max( "KEP_P_kWh")' in cell_text:
                    cell.value = today_max_kwh
                    
                # (참고) 22행 월간 전력량 표식도 영문으로 되어 있다면 함께 처리
                elif 'start("KEP_P_kWh")' in cell_text:
                    # 월시작 전력량 조회 로직 적용 (필요시)
                    pass
                elif 'end("KEP_P_kWh")' in cell_text:
                    cell.value = today_max_kwh  # 오늘 최종 지침과 동일
                    
    conn.close()


if __name__ == "__main__":
    generate_excel_report("2026-05-21")