# excel_report.py
import shutil
import sys
import os
import zipfile
import re
from datetime import datetime, timedelta
import openpyxl
# db_manager에서 MariaDB 연결 함수 및 주요 상수를 가져옵니다.
from db_manager import (DATA_LABELS, METER_FIELDS, DB_DIR, 
                        create_manual_meter_table, get_field_inspections_for_date,
                        get_db_raw_connection)

TEMPLATE_NAME = "template_전기실_운영일지.xlsx"
TEMPLATE_IN_APPDATA = os.path.join(DB_DIR, TEMPLATE_NAME)

def get_bundle_template_path(relative_path):
    """PyInstaller (.exe) 환경과 일반 파이썬 (.py) 환경을 모두 지원하는 경로 추적 함수"""
    try:
        base_path = sys._MEIPASS
        return os.path.join(base_path, relative_path)
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_path, relative_path)

def ensure_excel_template():
    """앱데이터 폴더에 템플릿 파일이 없으면 자동으로 원본을 찾아서 복사해 둡니다."""
    if not os.path.exists(TEMPLATE_IN_APPDATA):
        bundled_template = get_bundle_template_path(TEMPLATE_NAME)
        if os.path.exists(bundled_template):
            if not os.path.exists(DB_DIR):
                os.makedirs(DB_DIR)
            shutil.copy(bundled_template, TEMPLATE_IN_APPDATA)
            print(f"✅ 엑셀 템플릿 파일이 앱데이터에 복구되었습니다.")
        else:
            print(f"❌ 원본 템플릿 파일({TEMPLATE_NAME})을 찾을 수 없습니다.")

# 프로그램 시작 시점에 무조건 디버깅 체크 구동
ensure_excel_template()

def clean_external_links_physically(file_path):
    """엑셀 내부의 외부 링크 파편들을 물리적으로 세척하는 함수"""
    temp_file_path = file_path + ".tmp"
    try:
        with zipfile.ZipFile(file_path, 'r') as zin:
            with zipfile.ZipFile(temp_file_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    filename = item.filename
                    if "externalLinks/" in filename:
                        continue
                    if filename == "xl/_rels/workbook.xml.rels":
                        data = zin.read(filename).decode('utf-8')
                        data_clean = re.sub(r'<Relationship[^>]*Type="[^"]*externalLink"[^>]*/>', '', data)
                        zout.writestr(filename, data_clean.encode('utf-8'))
                        continue
                    if filename == "xl/workbook.xml":
                        data = zin.read(filename).decode('utf-8')
                        data_clean = re.sub(r'<externalReferences>.*?</externalReferences>', '', data, flags=re.DOTALL)
                        zout.writestr(filename, data_clean.encode('utf-8'))
                        continue
                    zout.writestr(item, zin.read(filename))
        os.remove(file_path)
        os.rename(temp_file_path, file_path)
    except Exception as e:
        if os.path.exists(temp_file_path): os.remove(temp_file_path)
        print(f"[경고] 외부 링크 물리 청소 중 예외 발생: {e}")

def generate_excel_report(selected_date, target_dir=None):
    """
    구역별 복잡한 수식 데이터 치환 엔진 (MariaDB 연동 버전)
    """
    template_file = TEMPLATE_IN_APPDATA

    if target_dir:
        output_file = os.path.join(target_dir, f"{selected_date}_전기실_운영일지.xlsx")
    else:  
        output_file = os.path.join(DB_DIR, f"{selected_date}_전기실_운영일지.xlsx")
    
    if not os.path.exists(TEMPLATE_IN_APPDATA):
        raise FileNotFoundError(f"템플릿 파일 [{template_file}]이 경로에 존재하지 않습니다.")
    
    shutil.copy(template_file, output_file)
    wb = openpyxl.load_workbook(output_file, data_only=False)
    
    if hasattr(wb, 'external_link_refs'): wb.external_link_refs = []
    if hasattr(wb, '_external_links'): wb._external_links = []

    dt = datetime.strptime(selected_date, "%Y-%m-%d")
    date_str_formatted = f"{dt.year}년 {dt.month:02d}월 {dt.day:02d}일"
    
    ws_summary = None
    ws_detail = None
    for sheet in wb.worksheets:
        if "운전현황" in sheet.title or "운영현황" in sheet.title:
            ws_summary = sheet
            sheet.title = f"{selected_date}_전력설비_운전현황"
        elif "상세내역" in sheet.title:
            ws_detail = sheet
            sheet.title = f"{selected_date}_상세내역"
            
    if not ws_summary: ws_summary = wb.worksheets[0]
    if not ws_detail: ws_detail = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]

    # 🌟 [MariaDB 전환] db_manager의 연결 빌더 활용
    conn = get_db_raw_connection()
    c = conn.cursor()

    # =========================================================================
    # [파트 1] 전력설비 운영 현황 분석 자료 - 최고/최저값 덮어쓰기 (상단 구역)
    # =========================================================================
    c.execute('SELECT * FROM daily_extremes WHERE log_date = %s', (selected_date,))
    rows_extreme = c.fetchall()
    
    # MariaDB 커서 필드명 추출
    columns_extreme = [desc[0] for desc in c.description] if c.description else []
    
    extremes_dict = {}
    for r in rows_extreme:
        ext_type = r[columns_extreme.index('extreme_type')]
        for col_name in DATA_LABELS:
            if col_name in columns_extreme:
                val = r[columns_extreme.index(col_name)]
                extremes_dict[(ext_type, col_name)] = round(float(val), 1) if val is not None else "-"

    for row in ws_summary.iter_rows(max_row=20):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_strip = cell.value.strip().replace(" ", "")
                if "일시:" in val_strip or ("일시" in val_strip and ":" in val_strip):
                    cell.value = f"일시 : {date_str_formatted}"
                elif "max(" in val_strip:
                    label = cell.value.split('"')[1].strip() if '"' in cell.value else cell.value.replace("max(", "").replace(")", "").strip()
                    cell.value = extremes_dict.get(("MAX", label), "-")
                elif "min(" in val_strip:
                    label = cell.value.split('"')[1].strip() if '"' in cell.value else cell.value.replace("min(", "").replace(")", "").strip()
                    cell.value = extremes_dict.get(("MIN", label), "-")

    # =========================================================================
    # [파트 2] 사용 전력량 현황 자동 데이터 치환 (21행 및 22행 영역)
    # =========================================================================
    day_min_val = "-"
    day_max_val = "-"
    if "KEP_P_kWh" in columns_extreme:
        day_min_val = extremes_dict.get(("MIN", "KEP_P_kWh"), "-")
        day_max_val = extremes_dict.get(("MAX", "KEP_P_kWh"), "-")
    
    if day_min_val == "-" or day_max_val == "-":
        c.execute('SELECT MIN(KEP_P_kWh), MAX(KEP_P_kWh) FROM raw_data WHERE log_date = %s', (selected_date,))
        res_raw = c.fetchone()
        if res_raw:
            if day_min_val == "-" and res_raw[0] is not None: day_min_val = round(float(res_raw[0]), 1)
            if day_max_val == "-" and res_raw[1] is not None: day_max_val = round(float(res_raw[1]), 1)

    start_of_month = f"{dt.year}-{dt.month:02d}-01"
    
    # 💡 MariaDB 호환 쿼리 (날짜 비교 범위 표준화 및 파라미터 %s 바인딩)
    c.execute('''
        SELECT KEP_P_kWh FROM raw_data 
        WHERE log_date >= %s AND log_date <= %s AND KEP_P_kWh IS NOT NULL 
        ORDER BY log_date ASC, log_time ASC LIMIT 1
    ''', (start_of_month, selected_date))
    month_start_res = c.fetchone()
    month_start_val = round(float(month_start_res[0]), 1) if month_start_res and month_start_res[0] is not None else "-"

    for row_idx in [21, 22]:
        for cell in ws_summary[row_idx]:
            if cell.value and isinstance(cell.value, str):
                text_clean = cell.value.strip().replace(" ", "")
                if row_idx == 21:
                    if 'min("KEP_P_kWh")' in text_clean: cell.value = day_min_val
                    elif 'max("KEP_P_kWh")' in text_clean: cell.value = day_max_val
                    elif 'E21-C21' in text_clean: cell.value = "=E21-C21"
                elif row_idx == 22:
                    if 'start("KEP_P_kWh")' in text_clean: cell.value = month_start_val
                    elif 'end("KEP_P_kWh")' in text_clean: cell.value = day_max_val
                    elif 'E22-C22' in text_clean: cell.value = "=E22-C22"

    # =========================================================================
    # [파트 3] 한전 메인 적산 사용량 하단 요약 테이블 연동 구역 (26행 이하)
    # =========================================================================
    today_mwh = day_max_val if day_max_val != "-" else "-"
    prev_mwh = day_min_val if day_min_val != "-" else "-"
    try:
        diff_mwh = round(float(today_mwh) - float(prev_mwh), 1) if today_mwh != "-" and prev_mwh != "-" else "-"
    except:
        diff_mwh = "-"

    for row in ws_summary.iter_rows(min_row=26, max_row=30):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_strip = cell.value.strip()
                if val_strip == "1469.79": cell.value = today_mwh
                elif val_strip == "1464.06": cell.value = prev_mwh
                elif val_strip == "5.7300000000000182": cell.value = diff_mwh

    # =========================================================================
    # [수동 검침] 수동 검침 데이터 연동 구역 (C27 ~ J35)
    # =========================================================================
    create_manual_meter_table()
    
    prev_day = (dt - timedelta(days=1)).strftime('%Y-%m-%d')
    first_day_of_current_month = dt.replace(day=1)
    last_day_of_prev_month_dt = first_day_of_current_month - timedelta(days=1)
    prev_month_last_day = last_day_of_prev_month_dt.strftime('%Y-%m-%d')
    
    meter_data = {"금일": {}, "전일": {}, "전월": {}}
    
    for label, target_d in [("금일", selected_date), ("전일", prev_day), ("전월", prev_month_last_day)]:
        fields_str = ", ".join(METER_FIELDS)
        # 💡 SQLite용 구문을 MariaDB 문법 파라미터(%s)로 수정
        c.execute(f'SELECT {fields_str} FROM manual_meter_logs WHERE log_date = %s', (target_d,))
        res = c.fetchone()
        
        if res:
            meter_data[label] = {field: res[idx] for idx, field in enumerate(METER_FIELDS)}
        else:
            meter_data[label] = {field: None for field in METER_FIELDS}

    for row in ws_summary.iter_rows(min_row=27, max_row=35, min_col=3, max_col=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_strip = cell.value.strip().replace(" ", "")
                
                match = re.match(r'(금일|전일|전월)\(["\']?([a-zA-Z0-9_]+)["\']?\)', val_strip)
                if match:
                    time_type = match.group(1)   
                    field_name = match.group(2)  
                    
                    if field_name in METER_FIELDS:
                        target_dict = meter_data.get(time_type, {})
                        db_val = target_dict.get(field_name, None)
                        cell.value = float(db_val) if db_val is not None else ""

    # =========================================================================
    # [파트 3-4] 현장 점검 데이터 자동 매핑 (E40:G48)
    # =========================================================================
    try:
        inspection_data = get_field_inspections_for_date(selected_date)
        for row in ws_summary["E40:G48"]:
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_text = cell.value.strip()
                    if "1차" in cell_text and "점검자" in cell_text:
                        cell.value = inspection_data[1]["name"] if inspection_data[1]["name"] else "-"
                    elif "1차" in cell_text and "시간" in cell_text:
                        cell.value = inspection_data[1]["time"] if inspection_data[1]["time"] else "-"
                    elif "2차" in cell_text and "점검자" in cell_text:
                        cell.value = inspection_data[2]["name"] if inspection_data[2]["name"] else "-"
                    elif "2차" in cell_text and "시간" in cell_text:
                        cell.value = inspection_data[2]["time"] if inspection_data[2]["time"] else "-"
                    elif "3차" in cell_text and "점검자" in cell_text:
                        cell.value = inspection_data[3]["name"] if inspection_data[3]["name"] else "-"
                    elif "3차" in cell_text and "시간" in cell_text:
                        cell.value = inspection_data[3]["time"] if inspection_data[3]["time"] else "-"
    except Exception as e:
        print(f"[ERROR] 현장 점검 서식 매핑 중 오류 발생: {e}")

    # =========================================================================
    # [파트 4] 상세내역 시트 - 24시간 시간별 데이터 주입 
    # =========================================================================
    for row in ws_detail.iter_rows(max_row=3):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "일자" in cell.value:
                cell.value = f"일자 : {date_str_formatted}"

    # 💡 MariaDB용 시간 추출 문법(HOUR()) 및 통계 그룹 쿼리로 변경
    avg_selects = ", ".join([f'AVG({name})' for name in DATA_LABELS])
    c.execute(f'''
        SELECT HOUR(log_time) as hour, {avg_selects} 
        FROM raw_data 
        WHERE log_date = %s 
        GROUP BY HOUR(log_time)
    ''', (selected_date,))
    hourly_rows = c.fetchall()
    
    hourly_data = {}
    for r in hourly_rows:
        hour = int(r[0]) if r[0] is not None else 0
        for idx, label in enumerate(DATA_LABELS):
            if r[idx + 1] is not None:
                hourly_data[(hour, label)] = round(float(r[idx + 1]), 1)

    for row_idx, row in enumerate(ws_detail.iter_rows(min_row=4, max_row=60), start=4):
        first_cell_val = str(row[0].value).strip() if row[0].value is not None else ""
        if first_cell_val == "0":
            col_mapping = {}
            for col_idx, cell in enumerate(row):
                if cell.value and isinstance(cell.value, str):
                    clean_label = cell.value.strip().replace('"', '')
                    if clean_label in DATA_LABELS:
                        col_mapping[col_idx] = clean_label
            
            if col_mapping:
                for h in range(24):
                    target_row = row_idx + h
                    ws_detail.cell(row=target_row, column=1).value = h
                    for col_idx, label in col_mapping.items():
                        val = hourly_data.get((h, label), "-")
                        ws_detail.cell(row=target_row, column=col_idx + 1).value = val

    c.close()
    conn.close()
    wb.save(output_file)
    
    clean_external_links_physically(output_file)
    print(f"✅ [출력 성공] 구역별 매핑 완료: {output_file}")

if __name__ == "__main__":
    generate_excel_report("2026-06-23")